from colorama import Fore, Back, init
import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd
from pathlib import Path
import time
from typing import Dict, List, Optional, Union
import json

# Web dependencies (optional, only used when RUN_WEB=1)
try:
    from flask import Flask, render_template, request, jsonify, abort, make_response
    from werkzeug.exceptions import BadRequest, NotFound, InternalServerError
    from werkzeug.utils import secure_filename
    import werkzeug
    try:
        from flask_cors import CORS
    except ImportError:
        CORS = None
except ImportError as e:
    Flask = None
    render_template = None
    CORS = None
    print(f"Flask dependencies not available: {e}")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def ID_kh():
    wb=load_workbook(filename="ThongTinKhachHang.xlsx")
    sheet=wb.active
    
    #Rut ra danh sach ma khach hang
    ma_kh_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            ma_kh_list.append(row[0])
            
    #tim ma khach hang lon nhat
    if not ma_kh_list:
        ID_kh = "DLT" + "0"*4 + "1"
        return ID_kh
    else:
        max_id = 0
        for ma_kh in ma_kh_list:
            try:
                num_part = int(ma_kh[3:])  # ma khach hang co dang DLT00001
                if num_part > max_id:
                    max_id = num_part
            except ValueError:
                continue  # Bỏ qua các mã không hợp lệ
        new_id_num = max_id + 1
        new_id = "DLU" + str(new_id_num).zfill(5)  # id luon co 5 chu so
        return new_id

def is_data_none():
    wb=load_workbook(filename="ThongTinKhachHang.xlsx")
    sheet=wb.active
    
    for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, values_only=True):
        if all(cell is None for cell in row):
            return True
        else:
            return False
        
def is_recycle_bin():
    # Dùng Path join để tránh lỗi escape chuỗi trên Windows
    file_path = Path("Recycle Bin") / "ThongTinKhachHang.xlsx"
    if file_path.exists() and file_path.is_file():
        return True

def check_file():
    if is_recycle_bin() == True: #False
        return False
    elif os.path.exists("ThongTinKhachHang.xlsx") == False:
        return False
    elif os.path.exists("ThongTinKhachHang.xlsx") and is_data_none() == True:
        return False
    else:
        return True
        

def show_customer_information():
    # Load excel file
    wb=load_workbook(filename="ThongTinKhachHang.xlsx")
    sheet=wb.active

    #create header from first row
    header = [cell.value for cell in sheet[1]]

    #create list of dictionaries from excel file
    data_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_rows.append(dict(zip(header, row)))
    #create dataframe from excel file
    df = pd.DataFrame(data_rows)
    return print(df)

def create_file():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Mã KH'
    ws['B1'] = 'Họ Tên'
    ws['C1'] = 'Số ĐT'
    ws['D1'] = 'Email'
    ws['E1'] = 'Địa Chỉ'
    
    wb.save("ThongTinKhachHang.xlsx")
    
def add_customer():
    new_customer = []
    #load excel file
    wb=load_workbook(filename="ThongTinKhachHang.xlsx")
    sheet=wb.active
    
    #get next row
    next_row = sheet.max_row + 1
    print(Fore.CYAN + Back.BLACK + "\n ----------THÊM KHÁCH HÀNG----------")
    ma_kh = ID_kh()
    ho_ten = input("Nhập họ tên khách hàng: ")
    so_dt = input("Nhập số điện thoại khách hàng: ")
    dia_chi = input("Nhập địa chỉ khách hàng: ")
    email = input("Nhập email khách hàng: ")
    new_customer.append(ma_kh)
    new_customer.append(ho_ten)
    new_customer.append(so_dt)
    new_customer.append(email)
    new_customer.append(dia_chi)
    
    # Xử lý trường Email không bắt buộc - nếu để trống thì gán khoảng trắng
    if new_customer[3] == "" or new_customer[3].strip() == "":
        new_customer[3] = " "
    Ten_truong = ["Mã KH", "Họ Tên", "Số ĐT", "Email", "Địa Chỉ"]
    
    while True:
        # Chỉ kiểm tra các trường bắt buộc: Họ Tên (index 1), Số ĐT (index 2), Địa Chỉ (index 4)
        required_fields = [1, 2, 4]  # Chỉ kiểm tra Họ Tên, Số ĐT, Địa Chỉ
        validation_passed = True
        
        for i in required_fields:
            if new_customer[i] == "" or new_customer[i].strip() == "":
                print(Fore.RED + Back.BLACK + f"Trường '{Ten_truong[i]}' không được để trống. Vui lòng nhập lại.")
                validation_passed = False
                if Ten_truong[i] == "Họ Tên":
                    ho_ten = input("Nhập họ tên khách hàng: ")
                    new_customer[i] = ho_ten
                elif Ten_truong[i] == "Số ĐT":
                    so_dt = input("Nhập số điện thoại khách hàng: ")
                    new_customer[i] = so_dt
                elif Ten_truong[i] == "Địa Chỉ":
                    dia_chi = input("Nhập địa chỉ khách hàng: ")
                    new_customer[i] = dia_chi
        
        if validation_passed:
            break
    for col, value in enumerate(new_customer, start=1):
        sheet.cell(row=next_row, column=col, value=str(value))
            
    wb.save("ThongTinKhachHang.xlsx")
    
    return print(Fore.GREEN + Back.BLACK + "Thêm khách hàng thành công.")
    
#########################
# Flask web integration #
#########################

def load_first_customer_for_web() -> dict | None:
    try:
        excel_path = Path("ThongTinKhachHang.xlsx")
        if not excel_path.exists():
            return None
        wb = load_workbook(filename=str(excel_path))
        sheet = wb.active
        header = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and any(cell is not None for cell in row):
                return dict(zip(header, row))
        return None
    except Exception:
        return None

def load_all_customers_for_web() -> list:
    """Load tất cả khách hàng từ Excel và phân nhóm"""
    try:
        excel_path = Path("ThongTinKhachHang.xlsx")
        if not excel_path.exists():
            return []
        
        wb = load_workbook(filename=str(excel_path))
        sheet = wb.active
        header = [cell.value for cell in sheet[1]]
        
        customers = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and any(cell is not None for cell in row):
                raw_data = dict(zip(header, row))
                
                # Chuẩn hóa dữ liệu
                customer = {
                    "code": raw_data.get("Mã KH") or raw_data.get("Ma KH") or raw_data.get("Code") or "",
                    "name": raw_data.get("Họ Tên") or raw_data.get("Ho Ten") or raw_data.get("Ten") or "",
                    "phone": raw_data.get("Số ĐT") or raw_data.get("So DT") or raw_data.get("SĐT") or "",
                    "email": raw_data.get("Email") or "",
                    "address": raw_data.get("Địa Chỉ") or raw_data.get("Dia Chi") or "",
                    "total_amount": raw_data.get("Tổng tiền mua") or raw_data.get("Tong Tien Mua") or "0đ",
                    "last_purchase": raw_data.get("Ngày cuối mua") or raw_data.get("Ngay Cuoi Mua") or "",
                }
                
                # Phân nhóm dựa trên tổng tiền mua
                total_amount_str = str(customer["total_amount"]).replace("đ", "").replace(",", "")
                try:
                    total_amount = float(total_amount_str) if total_amount_str else 0
                except ValueError:
                    total_amount = 0
                
                if total_amount >= 10000000:  # >= 10 triệu
                    customer["group"] = "vip"
                elif total_amount >= 5000000:  # >= 5 triệu
                    customer["group"] = "loyal"
                else:
                    customer["group"] = "potential"
                
                # Phân trạng thái Active/Inactive dựa trên lần cuối mua hàng
                last_purchase = customer["last_purchase"]
                if last_purchase and last_purchase != "Chưa có":
                    # Giả sử nếu có mua hàng trong 6 tháng gần đây thì Active
                    customer["status"] = "active"
                else:
                    customer["status"] = "inactive"
                
                customers.append(customer)
        
        # Sắp xếp theo alphabet dựa trên chữ cái cuối của tên
        customers.sort(key=lambda x: x["name"][-1].lower() if x["name"] else "z")
        
        return customers
    except Exception:
        return []

def load_customer_by_code_for_web(code: str) -> dict | None:
    """Tải 1 khách hàng theo mã từ Excel (phục vụ trang web)."""
    try:
        if not code:
            return None
        excel_path = Path("ThongTinKhachHang.xlsx")
        if not excel_path.exists():
            return None

        wb = load_workbook(filename=str(excel_path))
        sheet = wb.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None for cell in row):
                continue
            raw = dict(zip(header, row))
            row_code = (
                raw.get("Mã KH")
                or raw.get("Ma KH")
                or raw.get("Code")
                or ""
            )
            if str(row_code) == str(code):
                customer = {
                    "name": raw.get("Họ Tên") or raw.get("Ho Ten") or raw.get("Ten") or "",
                    "dob": raw.get("Ngày sinh") or raw.get("Ngay Sinh") or "",
                    "phone": raw.get("Số ĐT") or raw.get("So DT") or raw.get("SĐT") or raw.get("Phone") or "",
                    "email": raw.get("Email") or "",
                    "address": raw.get("Địa Chỉ") or raw.get("Dia Chi") or raw.get("Address") or "",
                    "code": row_code,
                    "total": raw.get("Tổng tiền mua") or raw.get("Tong Tien Mua") or "0đ",
                    "last_purchase": raw.get("Ngày cuối mua") or raw.get("Ngay Cuoi Mua") or "",
                }

                # Xác định group/status giống logic danh sách
                total_amount_str = str(customer["total"]).replace("đ", "").replace(",", "")
                try:
                    total_amount = float(total_amount_str) if total_amount_str else 0
                except ValueError:
                    total_amount = 0

                if total_amount >= 10000000:
                    customer["group"] = "vip"
                elif total_amount >= 5000000:
                    customer["group"] = "loyal"
                else:
                    customer["group"] = "potential"

                last_purchase = customer["last_purchase"]
                customer["status"] = "active" if last_purchase and last_purchase != "Chưa có" else "inactive"

                return customer
        return None
    except Exception:
        return None

def get_customer_stats(customers: list) -> dict:
    """Tính toán thống kê khách hàng"""
    active_customers = len([c for c in customers if c["status"] == "active"])
    
    # Tính tổng tiền của khách hàng Active
    active_amount = 0
    for customer in customers:
        if customer["status"] == "active":
            total_amount_str = str(customer["total_amount"]).replace("đ", "").replace(",", "")
            try:
                active_amount += float(total_amount_str) if total_amount_str else 0
            except ValueError:
                pass
    
    # Giả sử số đơn hàng hoạt động = số khách hàng active * 2 (trung bình)
    active_orders = active_customers * 2
    
    stats = {
        "total_customers": len(customers),
        "active_customers": active_customers,
        "active_orders": active_orders,
        "active_amount": f"{active_amount:,.0f}đ" if active_amount > 0 else "0đ",
    }
    return stats

# Create Flask app only if Flask is available
app = None
if Flask:
    base_dir = Path(__file__).resolve().parent
    template_dir = base_dir / "templates"
    static_dir = base_dir / "static"
    
    app = Flask(
        __name__,
        template_folder=str(template_dir),
        static_folder=str(static_dir),
    )
    
    # Configure Flask app
    app.config.update(
        SECRET_KEY=os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production'),
        MAX_CONTENT_LENGTH=16 * 1024 * 1024,  # 16MB max file size
        UPLOAD_FOLDER=str(base_dir / "uploads"),
        JSON_SORT_KEYS=False,
        JSONIFY_PRETTYPRINT_REGULAR=True
    )
    
    # Enable CORS for API endpoints
    if CORS:
        CORS(app, resources={
            r"/api/*": {"origins": "*"},
            r"/customer-dashboard": {"origins": "*"},
            r"/customers-list": {"origins": "*"}
        })
    
    # Create uploads directory if it doesn't exist
    upload_dir = base_dir / "uploads"
    upload_dir.mkdir(exist_ok=True)
    
    # Error handlers
    @app.errorhandler(400)
    def bad_request(error):
        logger.warning(f"Bad request: {error}")
        return jsonify({'error': 'Bad request', 'message': str(error)}), 400
    
    @app.errorhandler(404)
    def not_found(error):
        logger.warning(f"Not found: {error}")
        return jsonify({'error': 'Not found', 'message': 'Resource not found'}), 404
    
    @app.errorhandler(500)
    def internal_error(error):
        logger.error(f"Internal server error: {error}")
        return jsonify({'error': 'Internal server error', 'message': 'Something went wrong'}), 500
    
    @app.errorhandler(Exception)
    def handle_exception(e):
        logger.error(f"Unhandled exception: {e}")
        return jsonify({'error': 'Internal server error', 'message': 'An unexpected error occurred'}), 500

if app:
    @app.route("/")
    def dashboard_index():
        return render_template("index.html")

    @app.route("/api/dashboard-stats")
    def api_dashboard_stats():
        """API endpoint trả về thống kê KPI từ Excel"""
        try:
            from flask import jsonify
            customers = load_all_customers_for_web()
            
            # Tính toán KPI
            total_revenue = 0
            total_orders = len(customers)
            total_customers = len(customers)
            active_customers = 0
            
            for customer in customers:
                total_str = str(customer.get("total_amount", "0")).replace("đ", "").replace(",", "")
                try:
                    amount = float(total_str) if total_str else 0
                    total_revenue += amount
                except ValueError:
                    pass
                
                if customer.get("status") == "active":
                    active_customers += 1
            
            # Giả định lợi nhuận gộp = 30% doanh thu
            gross_profit = total_revenue * 0.3
            
            # AOV (Average Order Value)
            aov = total_revenue / total_orders if total_orders > 0 else 0
            
            # Conversion rate giả định 3.8%
            conversion_rate = 3.8
            
            # Hàng tồn kho giá trị cao (giả định)
            high_value_stock = 92000000
            
            # Top sản phẩm (giả định từ dữ liệu khách hàng)
            top_products = [
                {"name": "Áo thun Basic", "revenue": 220000000, "profit": 66000000, "quantity": 1200, "margin": 30},
                {"name": "Quần jeans Slim", "revenue": 180000000, "profit": 54000000, "quantity": 760, "margin": 30},
                {"name": "Giày Runner Pro", "revenue": 150000000, "profit": 60000000, "quantity": 310, "margin": 40},
                {"name": "Mũ lưỡi trai", "revenue": 90000000, "profit": 22500000, "quantity": 900, "margin": 25},
                {"name": "Túi đeo chéo", "revenue": 75000000, "profit": 26250000, "quantity": 215, "margin": 35},
            ]
            
            # Top khách hàng (lấy từ dữ liệu thật)
            top_customers_data = sorted(customers, key=lambda x: float(str(x.get("total_amount", "0")).replace("đ", "").replace(",", "") or 0), reverse=True)[:3]
            top_customers = []
            for c in top_customers_data:
                top_customers.append({
                    "name": c.get("name", ""),
                    "lifetime_value": c.get("total_amount", "0đ"),
                    "orders": 10,  # giả định
                    "last_purchase": c.get("last_purchase", "Chưa có")
                })
            
            # Dữ liệu biểu đồ doanh thu theo tuần
            revenue_chart = {
                "labels": ["Tuần 1", "Tuần 2", "Tuần 3", "Tuần 4"],
                "revenue": [280000000, 320000000, 260000000, 385000000],
                "profit": [90000000, 110000000, 95000000, 130000000]
            }
            
            # Cơ cấu sản phẩm
            product_mix = {
                "labels": ["Áo", "Quần", "Giày", "Phụ kiện"],
                "data": [38, 26, 22, 14]
            }
            
            # Doanh số theo kênh
            sales_by_channel = {
                "labels": ["Website", "Facebook", "Shopee", "Cửa hàng"],
                "data": [180000000, 95000000, 120000000, 140000000]
            }
            
            return jsonify({
                "kpi": {
                    "total_revenue": total_revenue,
                    "gross_profit": gross_profit,
                    "total_orders": total_orders,
                    "conversion_rate": conversion_rate,
                    "aov": aov,
                    "high_value_stock": high_value_stock
                },
                "top_products": top_products,
                "top_customers": top_customers,
                "charts": {
                    "revenue": revenue_chart,
                    "product_mix": product_mix,
                    "sales_by_channel": sales_by_channel
                }
            })
        except Exception as e:
            from flask import jsonify
            return jsonify({"error": str(e)}), 500

    @app.route("/customer-dashboard")
    def customer_dashboard():
        code = request.args.get("code") if request else None
        customer = load_customer_by_code_for_web(code) if code else None
        if not customer:
            raw = load_first_customer_for_web()
            if raw:
                customer = {
                    "name": raw.get("Họ Tên") or raw.get("Ho Ten") or raw.get("Ten") or "",
                    "dob": raw.get("Ngày sinh") or raw.get("Ngay Sinh") or "",
                    "phone": raw.get("Số ĐT") or raw.get("So DT") or raw.get("SĐT") or raw.get("Phone") or "",
                    "email": raw.get("Email") or "",
                    "address": raw.get("Địa Chỉ") or raw.get("Dia Chi") or raw.get("Address") or "",
                    "code": raw.get("Mã KH") or raw.get("Ma KH") or raw.get("Code") or "",
                    "total": raw.get("Tổng tiền mua") or raw.get("Tong Tien Mua") or "0đ",
                    "last_purchase": raw.get("Ngày cuối mua") or raw.get("Ngay Cuoi Mua") or "",
                }

                # Bổ sung group/status để đồng bộ với UI
                total_amount_str = str(customer["total"]).replace("đ", "").replace(",", "")
                try:
                    total_amount = float(total_amount_str) if total_amount_str else 0
                except ValueError:
                    total_amount = 0
                if total_amount >= 10000000:
                    customer["group"] = "vip"
                elif total_amount >= 5000000:
                    customer["group"] = "loyal"
                else:
                    customer["group"] = "potential"
                last_purchase = customer["last_purchase"]
                customer["status"] = "active" if last_purchase and last_purchase != "Chưa có" else "inactive"

        sample_orders = [
            {
                "code": "DLU00001",
                "customer": (customer["name"] if customer and customer.get("name") else "Nguyễn Phước Lộc"),
                "export_status": "Đã xuất",
                "value": "1,500,000đ",
                "date": "28/09/2025",
                "status": "done",
                "status_label": "Hoàn thành",
            },
            {
                "code": "DLU00002",
                "customer": (customer["name"] if customer and customer.get("name") else "Nguyễn Phước Lộc"),
                "export_status": "Chưa xuất",
                "value": "2,000,000đ",
                "date": "29/09/2025",
                "status": "processing",
                "status_label": "Đang xử lý",
            },
        ]

        return render_template(
            "customer_dashboard.html",
            active="customers",
            user_name="Admin",
            customer=customer,
            orders=sample_orders,
        )

    @app.route("/customers-list")
    def customers_list():
        customers = load_all_customers_for_web()
        stats = get_customer_stats(customers)
        
        return render_template(
            "customer_list.html",
            active="customers",
            customers=customers,
            stats=stats,
        )

    @app.route("/orders")
    def orders_page():
        return render_template("orders.html", active="orders")

    @app.route("/products")
    def products_page():
        return render_template("products.html", active="products")

    @app.route("/reports")
    def reports_page():
        return render_template("reports.html", active="reports")

    @app.route("/statistics")
    def statistics_page():
        return render_template("statistics.html", active="statistics")
    
    # API Routes for CRUD operations
    @app.route("/api/customers", methods=["GET"])
    def api_get_customers():
        """API endpoint để lấy danh sách tất cả khách hàng"""
        try:
            customers = load_all_customers_for_web()
            return jsonify({
                "success": True,
                "data": customers,
                "count": len(customers)
            })
        except Exception as e:
            logger.error(f"Error getting customers: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route("/api/customers/<customer_code>", methods=["GET"])
    def api_get_customer(customer_code):
        """API endpoint để lấy thông tin một khách hàng theo mã"""
        try:
            customer = load_customer_by_code_for_web(customer_code)
            if not customer:
                return jsonify({"success": False, "error": "Customer not found"}), 404
            return jsonify({"success": True, "data": customer})
        except Exception as e:
            logger.error(f"Error getting customer {customer_code}: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route("/api/customers", methods=["POST"])
    def api_create_customer():
        """API endpoint để tạo khách hàng mới"""
        try:
            from validators import validate_and_sanitize_customer
            
            data = request.get_json()
            if not data:
                return jsonify({"success": False, "error": "No data provided"}), 400
            
            # Validate and sanitize data
            sanitized_data, errors = validate_and_sanitize_customer(data)
            if errors:
                return jsonify({
                    "success": False, 
                    "error": "Validation failed", 
                    "details": errors
                }), 400
            
            # Generate new customer ID
            new_id = ID_kh()
            
            # Load Excel file
            wb = load_workbook(filename="ThongTinKhachHang.xlsx")
            sheet = wb.active
            next_row = sheet.max_row + 1
            
            # Add new customer data
            customer_data = [
                new_id,
                sanitized_data["name"],
                sanitized_data["phone"],
                sanitized_data["email"],
                sanitized_data["address"]
            ]
            
            for col, value in enumerate(customer_data, start=1):
                sheet.cell(row=next_row, column=col, value=str(value))
            
            wb.save("ThongTinKhachHang.xlsx")
            
            logger.info(f"Created new customer: {new_id}")
            return jsonify({
                "success": True,
                "message": "Customer created successfully",
                "data": {"code": new_id, **sanitized_data}
            }), 201
            
        except Exception as e:
            logger.error(f"Error creating customer: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route("/api/customers/<customer_code>", methods=["PUT"])
    def api_update_customer(customer_code):
        """API endpoint để cập nhật thông tin khách hàng"""
        try:
            from validators import validate_and_sanitize_customer
            
            data = request.get_json()
            if not data:
                return jsonify({"success": False, "error": "No data provided"}), 400
            
            # Validate and sanitize data
            sanitized_data, errors = validate_and_sanitize_customer(data)
            if errors:
                return jsonify({
                    "success": False, 
                    "error": "Validation failed", 
                    "details": errors
                }), 400
            
            # Load Excel file
            wb = load_workbook(filename="ThongTinKhachHang.xlsx")
            sheet = wb.active
            header = [cell.value for cell in sheet[1]]
            
            # Find customer row
            customer_row = None
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row and row[0] == customer_code:
                    customer_row = row_num
                    break
            
            if not customer_row:
                return jsonify({"success": False, "error": "Customer not found"}), 404
            
            # Update customer data
            field_mapping = {
                "name": "Họ Tên",
                "phone": "Số ĐT", 
                "email": "Email",
                "address": "Địa Chỉ"
            }
            
            for field, excel_col in field_mapping.items():
                if field in sanitized_data:
                    col_index = header.index(excel_col) + 1
                    sheet.cell(row=customer_row, column=col_index, value=str(sanitized_data[field]))
            
            wb.save("ThongTinKhachHang.xlsx")
            
            logger.info(f"Updated customer: {customer_code}")
            return jsonify({
                "success": True,
                "message": "Customer updated successfully"
            })
            
        except Exception as e:
            logger.error(f"Error updating customer {customer_code}: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route("/api/customers/<customer_code>", methods=["DELETE"])
    def api_delete_customer(customer_code):
        """API endpoint để xóa khách hàng"""
        try:
            # Load Excel file
            wb = load_workbook(filename="ThongTinKhachHang.xlsx")
            sheet = wb.active
            
            # Find customer row
            customer_row = None
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row and row[0] == customer_code:
                    customer_row = row_num
                    break
            
            if not customer_row:
                return jsonify({"success": False, "error": "Customer not found"}), 404
            
            # Delete customer row
            sheet.delete_rows(customer_row)
            wb.save("ThongTinKhachHang.xlsx")
            
            logger.info(f"Deleted customer: {customer_code}")
            return jsonify({
                "success": True,
                "message": "Customer deleted successfully"
            })
            
        except Exception as e:
            logger.error(f"Error deleting customer {customer_code}: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
    
    @app.route("/api/health")
    def api_health():
        """Health check endpoint"""
        return jsonify({
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "version": "1.0.0"
        })
    
    @app.route("/api/export/customers")
    def api_export_customers():
        """API endpoint để export danh sách khách hàng"""
        try:
            customers = load_all_customers_for_web()
            
            # Create response with Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Danh sách khách hàng"
            
            # Headers
            headers = ["Mã KH", "Họ Tên", "Số ĐT", "Email", "Địa Chỉ", "Tổng tiền mua", "Ngày cuối mua"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            for row, customer in enumerate(customers, 2):
                ws.cell(row=row, column=1, value=customer.get("code", ""))
                ws.cell(row=row, column=2, value=customer.get("name", ""))
                ws.cell(row=row, column=3, value=customer.get("phone", ""))
                ws.cell(row=row, column=4, value=customer.get("email", ""))
                ws.cell(row=row, column=5, value=customer.get("address", ""))
                ws.cell(row=row, column=6, value=customer.get("total_amount", ""))
                ws.cell(row=row, column=7, value=customer.get("last_purchase", ""))
            
            # Save to response
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting customers: {e}")
            return jsonify({"success": False, "error": str(e)}), 500

##########################
# HAM CHINH CHUONG TRINH #
##########################
def main():
    
    init(autoreset=False)
    print(Fore.YELLOW + Back.BLACK + "\n ----------QUẢN LÝ KHÁCH HÀNG----------")
    print("1. Thêm khách hàng")
    print("2. Hiển thị khách hàng")
    print("3. Tìm kiếm khách hàng")
    print("4. Cập nhật khách hàng")
    print("5. Xóa khách hàng")
    print("0. Thoát")
    print(Fore.CYAN + Back.BLACK)
    choice = input("Nhập lựa chọn của bạn: ")
    
    if choice == '1':
        if not os.path.exists("ThongTinKhachHang.xlsx"):
            print(Fore.YELLOW + Back.BLACK + "Chưa có file dữ liệu khách hàng. Đang tạo file...")
            time.sleep(1)
            create_file()
            print(Fore.GREEN + Back.BLACK + "Tạo file thành công.")
        elif os.path.exists("ThongTinKhachHang.xlsx") and is_data_none() == True:
            print(Fore.YELLOW + Back.BLACK + "File dữ liệu khách hàng hiện tại đang trống. Vui lòng thêm khách hàng.")
            add_customer()
        else:
            add_customer()
    elif choice == '2':
        if check_file() == False:
            print(Fore.RED + Back.BLACK + "Chưa có dữ liệu khách hàng. Vui lòng thêm khách hàng trước!!!")
        else:
            show_customer_information()
    elif choice == '3':
        if not os.path.exists("ThongTinKhachHang.xlsx"):
            print(Fore.YELLOW + Back.BLACK + "Chưa có file dữ liệu khách hàng. Đang tạo file...")
            time.sleep(1)
            create_file()
            print(Fore.GREEN + Back.BLACK + "Tạo file thành công.")
        elif os.path.exists("ThongTinKhachHang.xlsx") and is_data_none() == True:
            print(Fore.YELLOW + Back.BLACK + "File dữ liệu khách hàng hiện tại đang trống. Vui lòng thêm khách hàng.")
    elif choice == '4':
        pass
    elif choice == '5':
        pass
    elif choice == '0':
        print(Fore.WHITE + Back.BLACK + "Thoát chương trình.")
        exit()
    else: print(Fore.RED + Back.BLACK + "Lựa chọn không hợp lệ. Vui lòng thử lại!!!")
    
if __name__ == "__main__":
    # Set RUN_WEB=1 to start Flask server; otherwise run CLI loop as before
    if os.getenv("RUN_WEB") == "1" and app is not None:
        logger.info("Starting Flask application...")
        app.run(
            debug=os.getenv("FLASK_DEBUG", "True").lower() == "true",
            host=os.getenv("FLASK_HOST", "0.0.0.0"),
            port=int(os.getenv("FLASK_PORT", "5000"))
        )
    else:
        logger.info("Starting CLI application...")
        while True:
            try:
                main()
            except KeyboardInterrupt:
                logger.info("Application interrupted by user")
                print(Fore.YELLOW + Back.BLACK + "\nTạm biệt!")
                break
            except Exception as e:
                logger.error(f"Unexpected error in main loop: {e}")
                print(Fore.RED + Back.BLACK + f"Lỗi không mong muốn: {e}")
                print("Vui lòng thử lại...")
                time.sleep(1)
