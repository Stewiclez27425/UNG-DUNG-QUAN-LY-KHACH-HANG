#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#HỆ THỐNG ĐĂNG NHẬP ĐƠN GIẢN CHO ỨNG DỤNG QUẢN LÝ KHÁCH HÀNG
#===========================================================
#Sử dụng functions thuần túy, không sử dụng OOP (Object-Oriented Programming)
#Lưu trữ thông tin người dùng trong file Excel userdatalogin.xlsx

# ===========================
# IMPORT CÁC THƯ VIỆN CẦN THIẾT
# ===========================
import os
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from functools import wraps

# Import các thư viện cần thiết cho Excel và Flask
try:
    from openpyxl import load_workbook, Workbook
    from flask import session, request, redirect, url_for, flash, render_template, jsonify
except ImportError as e:
    print(f"❌ Lỗi import thư viện: {e}")
    print("📦 Vui lòng cài đặt các thư viện cần thiết:")
    print("   pip install openpyxl flask")
    print("   hoặc: pip install -r requirements.txt")

# ===========================
# BIẾN TOÀN CỤC VÀ CẤU HÌNH
# ===========================

# Tên file Excel chứa dữ liệu người dùng
USERDATA_FILE = "userdatalogin.xlsx"

# Dictionary lưu trữ thông tin tất cả người dùng trong bộ nhớ
# Cấu trúc: {username: {thông tin người dùng}}
users_data = {}

# Dictionary định nghĩa các cấp độ quyền truy cập
# Số càng nhỏ thì quyền càng cao
PERMISSION_LEVELS = {
    "FULL": 1,          # Quyền đầy đủ (Admin)
    "READ_WRITE": 2,    # Quyền đọc và ghi (Manager)
    "READ_ONLY": 3      # Chỉ quyền đọc (Staff)
}

# ===========================
# CÁC FUNCTIONS XỬ LÝ FILE EXCEL
# ===========================

def create_userdata_file():
    """
    TẠO FILE EXCEL CHỨA DỮ LIỆU NGƯỜI DÙNG
    =====================================
    
    Chức năng:
    - Tạo file Excel mới nếu chưa tồn tại
    - Thiết lập cấu trúc bảng với các cột cần thiết
    - Thêm dữ liệu mẫu cho các tài khoản mặc định
    
    Cấu trúc file Excel:
    - Cột A: Username (Tên đăng nhập)
    - Cột B: Password (Mật khẩu)
    - Cột C: FullName (Họ tên đầy đủ)
    - Cột D: Level (Cấp bậc: 1=Admin, 2=Manager, 3=Staff)
    - Cột E: Permission (Quyền: FULL/READ_WRITE/READ_ONLY)
    - Cột F: Status (Trạng thái: ACTIVE/INACTIVE)
    - Cột G: CreatedDate (Ngày tạo tài khoản)
    
    Returns:
        bool: True nếu tạo file thành công, False nếu có lỗi
    """
    try:
        # Kiểm tra xem file đã tồn tại chưa
        if os.path.exists(USERDATA_FILE):
            print(f"📁 File {USERDATA_FILE} đã tồn tại, không cần tạo mới")
            return True
        
        print(f"🔨 Đang tạo file {USERDATA_FILE}...")
        
        # Tạo workbook và worksheet mới
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        
        # Định nghĩa các cột header
        headers = [
            "Username",     # Tên đăng nhập (duy nhất)
            "Password",     # Mật khẩu (plain text - trong thực tế nên mã hóa)
            "FullName",     # Họ tên đầy đủ
            "Level",        # Cấp bậc (1-3)
            "Permission",   # Quyền truy cập
            "Status",       # Trạng thái tài khoản
            "CreatedDate"   # Ngày tạo
        ]
        
        # Ghi header vào dòng đầu tiên
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Dữ liệu mẫu cho các tài khoản mặc định
        sample_data = [
            # [Username, Password, FullName, Level, Permission, Status, CreatedDate]
            ["admin", "123456", "Nguyen Van Admin", 1, "FULL", "ACTIVE", "01/01/2024"],
            ["manager", "654321", "Tran Thi Manager", 2, "READ_WRITE", "ACTIVE", "01/01/2024"],
            ["staff1", "111111", "Le Van Staff", 3, "READ_ONLY", "ACTIVE", "01/01/2024"],
            ["staff2", "222222", "Pham Thi Staff", 3, "READ_ONLY", "ACTIVE", "01/01/2024"],
            ["demo", "demo123", "Demo User", 3, "READ_ONLY", "INACTIVE", "01/01/2024"]
        ]
        
        # Ghi dữ liệu mẫu vào file Excel
        for row_idx, data in enumerate(sample_data, 2):  # Bắt đầu từ dòng 2
            for col_idx, value in enumerate(data, 1):    # Bắt đầu từ cột 1
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Lưu file Excel
        wb.save(USERDATA_FILE)
        print(f"✅ Đã tạo file {USERDATA_FILE} thành công với {len(sample_data)} tài khoản mẫu")
        
        # In thông tin các tài khoản mẫu
        print("📋 Các tài khoản mẫu đã được tạo:")
        print("   👤 admin/123456 - Quyền FULL (Admin)")
        print("   👤 manager/654321 - Quyền READ_WRITE (Manager)")
        print("   👤 staff1/111111 - Quyền READ_ONLY (Staff)")
        print("   👤 staff2/222222 - Quyền READ_ONLY (Staff)")
        print("   👤 demo/demo123 - Tài khoản bị khóa")
        
        return True
        
    except Exception as e:
        print(f"❌ Lỗi khi tạo file {USERDATA_FILE}: {e}")
        return False

def load_users():
    """
    TẢI DỮ LIỆU NGƯỜI DÙNG TỪ FILE EXCEL VÀO BỘ NHỚ
    ===============================================
    
    Chức năng:
    - Đọc toàn bộ dữ liệu từ file Excel
    - Chuyển đổi thành dictionary để xử lý nhanh
    - Lưu vào biến toàn cục users_data
    - Tự động tạo file nếu chưa tồn tại
    
    Global Variables:
        users_data (dict): Dictionary chứa thông tin tất cả người dùng
    
    Returns:
        bool: True nếu tải thành công, False nếu có lỗi
    """
    global users_data
    
    try:
        print(f"📖 Đang tải dữ liệu người dùng từ {USERDATA_FILE}...")
        
        # Kiểm tra và tạo file nếu chưa tồn tại
        if not os.path.exists(USERDATA_FILE):
            print(f"⚠️  File {USERDATA_FILE} chưa tồn tại, đang tạo mới...")
            if not create_userdata_file():
                return False
        
        # Mở file Excel
        wb = load_workbook(USERDATA_FILE)
        ws = wb.active
        
        # Lấy danh sách header từ dòng đầu tiên
        header = [cell.value for cell in ws[1]]
        print(f"📊 Cấu trúc dữ liệu: {header}")
        
        # Khởi tạo lại dictionary users_data
        users_data = {}
        
        # Đọc từng dòng dữ liệu (bỏ qua dòng header)
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Kiểm tra dòng có dữ liệu không (không phải dòng trống)
            if row and any(cell is not None for cell in row):
                # Tạo dictionary cho một người dùng
                user_info = dict(zip(header, row))
                
                # Lấy username làm key
                username = user_info.get("Username", "")
                if username and username.strip():  # Kiểm tra username không rỗng
                    # Chuẩn hóa dữ liệu
                    user_info["Username"] = str(username).strip()
                    user_info["Password"] = str(user_info.get("Password", "")).strip()
                    user_info["FullName"] = str(user_info.get("FullName", "")).strip()
                    user_info["Level"] = int(user_info.get("Level", 3))
                    user_info["Permission"] = str(user_info.get("Permission", "READ_ONLY")).strip()
                    user_info["Status"] = str(user_info.get("Status", "INACTIVE")).strip()
                    user_info["CreatedDate"] = str(user_info.get("CreatedDate", "")).strip()
                    
                    # Lưu vào dictionary
                    users_data[username] = user_info
                    row_count += 1
        
        print(f"✅ Đã tải thành công {row_count} người dùng từ {USERDATA_FILE}")
        
        # In thống kê nhanh
        active_users = sum(1 for user in users_data.values() if user.get("Status") == "ACTIVE")
        print(f"📈 Thống kê: {active_users}/{row_count} tài khoản đang hoạt động")
        
        return True
        
    except Exception as e:
        print(f"❌ Lỗi khi tải dữ liệu người dùng: {e}")
        users_data = {}  # Reset về dictionary rỗng nếu có lỗi
        return False

def save_users():
    """
    LƯU DỮ LIỆU NGƯỜI DÙNG TỪ BỘ NHỚ VÀO FILE EXCEL
    ===============================================
    
    Chức năng:
    - Ghi toàn bộ dữ liệu từ users_data vào file Excel
    - Tạo file Excel mới (ghi đè file cũ)
    - Đảm bảo cấu trúc dữ liệu nhất quán
    
    Global Variables:
        users_data (dict): Dictionary chứa thông tin tất cả người dùng
    
    Returns:
        bool: True nếu lưu thành công, False nếu có lỗi
    """
    global users_data
    
    try:
        print(f"💾 Đang lưu {len(users_data)} người dùng vào {USERDATA_FILE}...")
        
        # Tạo workbook và worksheet mới
        wb = Workbook()
        ws = wb.active
        ws.title = "UserData"
        
        # Định nghĩa các cột header (phải giống với create_userdata_file)
        headers = ["Username", "Password", "FullName", "Level", "Permission", "Status", "CreatedDate"]
        
        # Ghi header vào dòng đầu tiên
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Ghi dữ liệu người dùng
        for row_idx, user_info in enumerate(users_data.values(), 2):  # Bắt đầu từ dòng 2
            ws.cell(row=row_idx, column=1, value=user_info.get("Username", ""))
            ws.cell(row=row_idx, column=2, value=user_info.get("Password", ""))
            ws.cell(row=row_idx, column=3, value=user_info.get("FullName", ""))
            ws.cell(row=row_idx, column=4, value=user_info.get("Level", 3))
            ws.cell(row=row_idx, column=5, value=user_info.get("Permission", "READ_ONLY"))
            ws.cell(row=row_idx, column=6, value=user_info.get("Status", "ACTIVE"))
            ws.cell(row=row_idx, column=7, value=user_info.get("CreatedDate", ""))
        
        # Lưu file Excel
        wb.save(USERDATA_FILE)
        print(f"✅ Đã lưu thành công {len(users_data)} người dùng vào {USERDATA_FILE}")
        
        return True
        
    except Exception as e:
        print(f"❌ Lỗi khi lưu dữ liệu người dùng: {e}")
        return False

# ===========================
# CÁC FUNCTIONS XÁC THỰC VÀ ĐĂNG NHẬP
# ===========================

def authenticate_user(username: str, password: str) -> Optional[Dict]:
    """
    XÁC THỰC ĐĂNG NHẬP NGƯỜI DÙNG
    =============================
    
    Chức năng:
    - Kiểm tra username và password có đúng không
    - Kiểm tra tài khoản có đang hoạt động không
    - Trả về thông tin người dùng nếu đăng nhập thành công
    
    Args:
        username (str): Tên đăng nhập
        password (str): Mật khẩu
    
    Returns:
        Optional[Dict]: Dictionary chứa thông tin người dùng nếu thành công, None nếu thất bại
    
    Example:
        user_info = authenticate_user("admin", "123456")
        if user_info:
            print(f"Đăng nhập thành công: {user_info['FullName']}")
        else:
            print("Đăng nhập thất bại!")
    """
    # Kiểm tra input
    if not username or not password:
        print("⚠️  Username và password không được để trống")
        return None
    
    # Kiểm tra username có tồn tại không
    if username not in users_data:
        print(f"❌ Không tìm thấy username '{username}'")
        return None
    
    user_info = users_data[username]
    
    # Kiểm tra password
    if user_info.get("Password") != password:
        print(f"❌ Mật khẩu không đúng cho username '{username}'")
        return None
    
    # Kiểm tra trạng thái tài khoản
    if user_info.get("Status") != "ACTIVE":
        print(f"❌ Tài khoản '{username}' đã bị khóa hoặc không hoạt động")
        return None
    
    # Đăng nhập thành công
    print(f"✅ Đăng nhập thành công: {user_info.get('FullName', username)}")
    return user_info

def get_user_info(username: str) -> Optional[Dict]:
    """
    LẤY THÔNG TIN CHI TIẾT CỦA NGƯỜI DÙNG
    ====================================
    
    Chức năng:
    - Trả về toàn bộ thông tin của một người dùng
    - Không yêu cầu xác thực mật khẩu
    
    Args:
        username (str): Tên đăng nhập
    
    Returns:
        Optional[Dict]: Dictionary chứa thông tin người dùng, None nếu không tìm thấy
    
    Example:
        user_info = get_user_info("admin")
        if user_info:
            print(f"Họ tên: {user_info['FullName']}")
            print(f"Quyền: {user_info['Permission']}")
    """
    return users_data.get(username)

def is_user_active(username: str) -> bool:
    """
    KIỂM TRA TRẠNG THÁI HOẠT ĐỘNG CỦA TÀI KHOẢN
    ==========================================
    
    Args:
        username (str): Tên đăng nhập
    
    Returns:
        bool: True nếu tài khoản đang hoạt động, False nếu không
    
    Example:
        if is_user_active("admin"):
            print("Tài khoản admin đang hoạt động")
        else:
            print("Tài khoản admin bị khóa")
    """
    user_info = users_data.get(username)
    if user_info:
        return user_info.get("Status") == "ACTIVE"
    return False

# ===========================
# CÁC FUNCTIONS KIỂM TRA QUYỀN TRUY CẬP
# ===========================

def check_permission(username: str, required_permission: str) -> bool:
    """
    KIỂM TRA QUYỀN TRUY CẬP CỦA NGƯỜI DÙNG
    =====================================
    
    Chức năng:
    - So sánh quyền của người dùng với quyền yêu cầu
    - Sử dụng hệ thống phân cấp: FULL > READ_WRITE > READ_ONLY
    
    Args:
        username (str): Tên đăng nhập
        required_permission (str): Quyền yêu cầu (FULL/READ_WRITE/READ_ONLY)
    
    Returns:
        bool: True nếu có quyền, False nếu không có quyền
    
    Example:
        if check_permission("admin", "FULL"):
            print("Admin có quyền đầy đủ")
        
        if check_permission("staff1", "READ_WRITE"):
            print("Staff có quyền ghi")  # Sẽ không in vì staff chỉ có READ_ONLY
    """
    # Kiểm tra user có tồn tại không
    user_info = users_data.get(username)
    if not user_info:
        return False
    
    # Kiểm tra tài khoản có hoạt động không
    if not is_user_active(username):
        return False
    
    # Lấy quyền của người dùng
    user_permission = user_info.get("Permission", "READ_ONLY")
    
    # Lấy level của quyền (số càng nhỏ thì quyền càng cao)
    user_level = PERMISSION_LEVELS.get(user_permission, 999)
    required_level = PERMISSION_LEVELS.get(required_permission, 999)
    
    # Kiểm tra quyền
    return user_level <= required_level

def check_user_level(username: str, min_level: int) -> bool:
    """
    KIỂM TRA CẤP BẬC TỐI THIỂU CỦA NGƯỜI DÙNG
    ========================================
    
    Chức năng:
    - Kiểm tra cấp bậc của người dùng có đủ điều kiện không
    - Cấp bậc: 1=Admin, 2=Manager, 3=Staff (số càng nhỏ thì cấp càng cao)
    
    Args:
        username (str): Tên đăng nhập
        min_level (int): Cấp bậc tối thiểu yêu cầu
    
    Returns:
        bool: True nếu đủ cấp bậc, False nếu không đủ
    
    Example:
        if check_user_level("admin", 1):
            print("Admin có cấp bậc đủ để truy cập")
        
        if check_user_level("staff1", 2):
            print("Staff có cấp Manager trở lên")  # Sẽ không in vì staff là cấp 3
    """
    # Kiểm tra user có tồn tại không
    user_info = users_data.get(username)
    if not user_info:
        return False
    
    # Kiểm tra tài khoản có hoạt động không
    if not is_user_active(username):
        return False
    
    # Lấy cấp bậc của người dùng
    user_level = user_info.get("Level", 999)
    
    # Kiểm tra cấp bậc (số càng nhỏ thì cấp càng cao)
    return user_level <= min_level

# ===========================
# CÁC FUNCTIONS QUẢN LÝ NGƯỜI DÙNG
# ===========================

def add_new_user(username: str, password: str, full_name: str, 
                 level: int, permission: str) -> bool:
    """
    THÊM NGƯỜI DÙNG MỚI VÀO HỆ THỐNG
    ===============================
    
    Chức năng:
    - Tạo tài khoản mới với thông tin đầy đủ
    - Kiểm tra username có bị trùng không
    - Tự động lưu vào file Excel
    
    Args:
        username (str): Tên đăng nhập (duy nhất)
        password (str): Mật khẩu
        full_name (str): Họ tên đầy đủ
        level (int): Cấp bậc (1=Admin, 2=Manager, 3=Staff)
        permission (str): Quyền truy cập (FULL/READ_WRITE/READ_ONLY)
    
    Returns:
        bool: True nếu thêm thành công, False nếu thất bại
    
    Example:
        success = add_new_user("newuser", "password123", "Nguyen Van New", 3, "READ_ONLY")
        if success:
            print("Đã tạo tài khoản thành công")
    """
    global users_data
    
    try:
        # Kiểm tra input
        if not username or not password or not full_name:
            print("❌ Username, password và họ tên không được để trống")
            return False
        
        # Kiểm tra username đã tồn tại chưa
        if username in users_data:
            print(f"❌ Username '{username}' đã tồn tại!")
            return False
        
        # Kiểm tra level hợp lệ
        if level not in [1, 2, 3]:
            print("❌ Cấp bậc phải là 1 (Admin), 2 (Manager), hoặc 3 (Staff)")
            return False
        
        # Kiểm tra permission hợp lệ
        if permission not in PERMISSION_LEVELS:
            print(f"❌ Quyền truy cập phải là một trong: {list(PERMISSION_LEVELS.keys())}")
            return False
        
        print(f"👤 Đang tạo tài khoản mới cho '{username}'...")
        
        # Tạo thông tin người dùng mới
        user_info = {
            "Username": username.strip(),
            "Password": password.strip(),
            "FullName": full_name.strip(),
            "Level": level,
            "Permission": permission.strip(),
            "Status": "ACTIVE",  # Mặc định là hoạt động
            "CreatedDate": datetime.now().strftime("%d/%m/%Y")
        }
        
        # Thêm vào dictionary
        users_data[username] = user_info
        
        # Lưu vào file Excel
        if save_users():
            print(f"✅ Đã tạo tài khoản '{username}' thành công!")
            print(f"   📋 Họ tên: {full_name}")
            print(f"   🎖️  Cấp bậc: {level}")
            print(f"   🔐 Quyền: {permission}")
            return True
        else:
            # Rollback nếu lưu file thất bại
            del users_data[username]
            print(f"❌ Không thể lưu tài khoản '{username}' vào file")
            return False
            
    except Exception as e:
        print(f"❌ Lỗi khi tạo tài khoản '{username}': {e}")
        return False

def print_all_users():
    """
    IN DANH SÁCH TẤT CẢ NGƯỜI DÙNG
    =============================
    """
    if not users_data:
        print("📋 Không có người dùng nào trong hệ thống")
        return
    
    print(f"\n📋 DANH SÁCH TẤT CẢ NGƯỜI DÙNG ({len(users_data)} người)")
    print("=" * 90)
    print(f"{'Username':<15} {'Họ Tên':<25} {'Cấp':<5} {'Quyền':<12} {'Trạng Thái':<12}")
    print("-" * 90)
    
    for user_info in users_data.values():
        status_icon = "🟢" if user_info.get("Status") == "ACTIVE" else "🔴"
        print(f"{user_info.get('Username', ''):<15} "
              f"{user_info.get('FullName', ''):<25} "
              f"{user_info.get('Level', ''):<5} "
              f"{user_info.get('Permission', ''):<12} "
              f"{status_icon} {user_info.get('Status', ''):<10}")

# ===========================
# CÁC FUNCTIONS DECORATOR CHO FLASK
# ===========================

def require_permission(permission: str):
    """
    DECORATOR KIỂM TRA QUYỀN TRUY CẬP CHO FLASK ROUTES
    ================================================
    
    Sử dụng với Flask để bảo vệ các route cần quyền đặc biệt
    
    Args:
        permission (str): Quyền yêu cầu
    
    Example:
        @app.route('/admin')
        @require_permission('FULL')
        def admin_page():
            return "Trang admin"
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Kiểm tra xem có thông tin user trong session không
            if 'username' not in session:
                flash('Vui lòng đăng nhập để truy cập!', 'error')
                return redirect(url_for('login'))
            
            username = session['username']
            
            # Kiểm tra quyền
            if not check_permission(username, permission):
                flash('Bạn không có quyền truy cập chức năng này!', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def require_level(min_level: int):
    """
    DECORATOR KIỂM TRA CẤP BẬC TỐI THIỂU CHO FLASK ROUTES
    ==================================================
    
    Args:
        min_level (int): Cấp bậc tối thiểu yêu cầu
    
    Example:
        @app.route('/manager')
        @require_level(2)
        def manager_page():
            return "Trang manager"
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Kiểm tra xem có thông tin user trong session không
            if 'username' not in session:
                flash('Vui lòng đăng nhập để truy cập!', 'error')
                return redirect(url_for('login'))
            
            username = session['username']
            
            # Kiểm tra cấp bậc
            if not check_user_level(username, min_level):
                flash('Bạn không có quyền truy cập chức năng này!', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_user_by_id(user_id: str) -> Optional[Dict]:
    """
    HÀM HELPER ĐỂ LẤY USER THEO ID CHO FLASK-LOGIN
    =============================================
    
    Args:
        user_id (str): ID của user (thường là username)
    
    Returns:
        Optional[Dict]: Thông tin user nếu tìm thấy
    """
    return get_user_info(user_id)

# ===========================
# CÁC FUNCTIONS TEST VÀ DEMO
# ===========================

def test_simple_login_system():
    """
    TEST HỆ THỐNG ĐĂNG NHẬP ĐƠN GIẢN
    ===============================
    
    Chức năng:
    - Test tạo file Excel
    - Test tải dữ liệu
    - Test xác thực đăng nhập
    - Test phân quyền
    - Test quản lý người dùng
    """
    print("\n🧪 TESTING HỆ THỐNG ĐĂNG NHẬP ĐƠN GIẢN")
    print("=" * 60)
    
    # Test 1: Tạo file userdata
    print("1. 📁 Tạo file dữ liệu người dùng...")
    create_userdata_file()
    
    # Test 2: Tải dữ liệu
    print("\n2. 📖 Tải dữ liệu người dùng...")
    load_users()
    
    # Test 3: Test đăng nhập
    print("\n3. 🔐 Test xác thực đăng nhập...")
    test_cases = [
        ("admin", "123456", True),
        ("manager", "654321", True),
        ("staff1", "111111", True),
        ("admin", "wrongpass", False),
        ("nonexistent", "123456", False),
        ("demo", "demo123", False)  # Tài khoản bị khóa
    ]
    
    for username, password, expected in test_cases:
        user_info = authenticate_user(username, password)
        result = "✅ PASS" if (user_info is not None) == expected else "❌ FAIL"
        print(f"   {username}/{password}: {result}")
    
    # Test 4: Test phân quyền
    print("\n4. 🔑 Test phân quyền...")
    permission_tests = [
        ("admin", "FULL", True),
        ("admin", "READ_WRITE", True),
        ("admin", "READ_ONLY", True),
        ("manager", "FULL", False),
        ("manager", "READ_WRITE", True),
        ("manager", "READ_ONLY", True),
        ("staff1", "FULL", False),
        ("staff1", "READ_WRITE", False),
        ("staff1", "READ_ONLY", True)
    ]
    
    for username, permission, expected in permission_tests:
        result_check = check_permission(username, permission)
        result = "✅ PASS" if result_check == expected else "❌ FAIL"
        print(f"   {username} có quyền {permission}: {result}")
    
    # Test 5: Test cấp bậc
    print("\n5. 🎖️  Test cấp bậc...")
    level_tests = [
        ("admin", 1, True),
        ("admin", 2, True),
        ("manager", 1, False),
        ("manager", 2, True),
        ("staff1", 2, False),
        ("staff1", 3, True)
    ]
    
    for username, min_level, expected in level_tests:
        result_check = check_user_level(username, min_level)
        result = "✅ PASS" if result_check == expected else "❌ FAIL"
        print(f"   {username} có cấp >= {min_level}: {result}")
    
    # Test 6: Test thêm người dùng mới
    print("\n6. 👤 Test thêm người dùng mới...")
    success = add_new_user("testuser", "test123", "Test User", 3, "READ_ONLY")
    print(f"   Thêm user mới: {'✅ PASS' if success else '❌ FAIL'}")
    
    # Test 7: Hiển thị danh sách người dùng
    print("\n7. 📋 Hiển thị danh sách người dùng...")
    print_all_users()
    
    print("\n✅ Hoàn thành test hệ thống đăng nhập đơn giản!")

def main():
    """
    HÀM MAIN - ĐIỂM KHỞI ĐẦU CỦA CHƯƠNG TRÌNH
    ========================================
    
    Chức năng:
    - Khởi tạo hệ thống
    - Tải dữ liệu người dùng
    - Chạy test demo
    """
    print("KHỞI ĐỘNG HỆ THỐNG ĐĂNG NHẬP ĐƠN GIẢN")
    print("=" * 50)
    
    # Khởi tạo hệ thống
    print("Đang khởi tạo hệ thống...")
    load_users()
    
    # Chạy test
    test_simple_login_system()
    
    print("\n🎉 Hệ thống đã sẵn sàng sử dụng!")
    print("📚 Các functions chính:")
    print("   - authenticate_user(username, password)")
    print("   - check_permission(username, permission)")
    print("   - check_user_level(username, min_level)")
    print("   - add_new_user(username, password, full_name, level, permission)")
    print("   - get_user_info(username)")
    print("   - is_user_active(username)")

if __name__ == "__main__":
    main()
