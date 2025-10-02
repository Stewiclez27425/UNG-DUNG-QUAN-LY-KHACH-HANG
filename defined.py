from colorama import Fore, Back, init
import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd
from pathlib import Path
import time
from typing import Dict, List, Optional, Union
import json

#kiem tra trung lap data
def is_dupe_data(header, value):
    Workbook = load_workbook("ThongTinKhachHang.xlsx")
    ws = Workbook.active
    
    if header == "SĐT":
        for cell in ws['C'][1:]:
            if value == cell.value: return True
        else:
            return False
    elif header == "Email":
        for cell in ws['D'][1:]:
            if value == cell.value: return True
        else:
            return False

#Phan loai khach hang (WIP)
def phan_loai():
   # """Phân loại khách hàng mặc định"""
    return "Khách hàng tiềm năng"

#TrangThai khach hang
def trang_thai_KH():
    #"""Trạng thái khách hàng mặc định"""
    return "Hoạt động"
#xuat tong tien tu hoa don (WIP)

#Cap nhap khach hang
def update_customer():
    """Cập nhật thông tin khách hàng với error handling"""
    try:
        KH_information = []
        Workbook=load_workbook("ThongTinKhachHang.xlsx")
        ws=Workbook.active
        
        #tao dict
        col_index = {cell.value: cell.column for cell in ws[1]}
        #Update Thong tin khach hang
        ma_kh = input ("Nhập mã khách hàng cần cập nhật thông tin: ")
        
        customer_found = False
        #duyet tung phan tu trong ma khach hang
        for cell in ws['A'][1:]:
            if ma_kh == cell.value:
                customer_found = True
                print("---------- CẬP NHẬT KHÁCH HÀNG ----------")
                #Thay doi thong tin khach hang
                ho_ten = input("Nhập họ tên khách hàng: ")
                so_dt = input("Nhập số điện thoại khách hàng: ")
                dia_chi = input("Nhập địa chỉ khách hàng: ")
                email = input("Nhập email khách hàng: ")
                KH_information.append(ho_ten)
                KH_information.append(so_dt)
                KH_information.append(email)
                KH_information.append(dia_chi)
                Ten_truong = ['Họ tên', 'Số ĐT', 'Email', 'Địa chỉ']
                #Bo qua email neu de trong
                if KH_information[2] == '' or KH_information[2].strip() == '':
                    KH_information[2] = ' '

                #kiem tra xem ho ten, dia chi, so dth co bi trong khong
                while True:
                    # Chỉ kiểm tra các trường bắt buộc: Họ Tên (index 0), Số ĐT (index 1), Địa Chỉ (index 3)
                    required_fields = [0, 1, 3]  # Chỉ kiểm tra Họ Tên, Số ĐT, Địa Chỉ
                    validation_passed = True
                    
                    for idx in required_fields:
                        if KH_information[idx] == "" or KH_information[idx].strip() == "":
                            print(Fore.RED + Back.BLACK + f"Trường '{Ten_truong[idx]}' không được để trống. Vui lòng nhập lại.")
                            if Ten_truong[idx] == "Họ tên":
                                KH_information[0] = input("Nhập họ tên khách hàng: ")
                            elif Ten_truong[idx] == "Số ĐT":
                                KH_information[1] = input("Nhập số điện thoại khách hàng: ")
                            elif Ten_truong[idx] == "Địa chỉ":
                                KH_information[3] = input("Nhập địa chỉ khách hàng: ")
                            validation_passed = False
                    if validation_passed:
                        break
                
                # Cập nhật dữ liệu
                print(Fore.GREEN + Back.BLACK + "Cập nhật khách hàng thành công")
                row = cell.row
                ws.cell(row=row, column=col_index["Họ Tên"]).value = ho_ten
                ws.cell(row=row, column=col_index["SĐT"]).value = so_dt
                ws.cell(row=row, column=col_index["Email"]).value = email
                ws.cell(row=row, column=col_index["Địa chỉ"]).value = dia_chi
                break
        
        if not customer_found:
            print(Fore.RED + Back.BLACK + "Không tìm thấy khách hàng có mã {}".format(ma_kh))
            return False
        
        # Lưu file
        Workbook.save("ThongTinKhachHang.xlsx")
        return True
        
    except FileNotFoundError:
        print(Fore.RED + Back.BLACK + "Không tìm thấy file ThongTinKhachHang.xlsx")
        return False
    except Exception as e:
        print(Fore.RED + Back.BLACK + f"Lỗi không mong muốn: {e}")
        return False

#generate ID khach hang   
def ID_kh():
    """Tạo ID khách hàng mới với error handling"""
    try:
        wb=load_workbook(filename="ThongTinKhachHang.xlsx")
        sheet=wb.active
    except FileNotFoundError:
        print(Fore.RED + Back.BLACK + "Không tìm thấy file ThongTinKhachHang.xlsx")
        return "DLT00001"  # Return default ID if file not found
    except Exception as e:
        print(Fore.RED + Back.BLACK + f"Lỗi khi mở file Excel: {e}")
        return "DLT00001"
    
    #Rut ra danh sach ma khach hang
    ma_kh_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            ma_kh_list.append(row[0])
            
    #tim ma khach hang lon nhat
    if not ma_kh_list:
        ID_kh = "DLT" + "0"*4 + "1" #Neu khong co ID thi tao moi
        return ID_kh
    else:
        max_id = 0
        for ma_kh in ma_kh_list:
            try:
                num_part = int(ma_kh[3:])  # ma khach hang co dang DLT00001
                if num_part > max_id:
                    max_id = num_part
            except ValueError:
                continue  # Bỏ qua các mã không hợp lệ
        new_id_num = max_id + 1
        new_id = "DLT" + str(new_id_num).zfill(5)  # id luon co 5 chu so
        return new_id
    
def is_data_none(Ten_File): #Kiem tra du lieu co bi rong khong
    """Kiểm tra dữ liệu có rỗng không với error handling"""
    try:
        wb=load_workbook(filename=Ten_File)
        sheet=wb.active
    except FileNotFoundError:
        return True  # File không tồn tại = dữ liệu rỗng
    except Exception as e:
        print(Fore.RED + Back.BLACK + f"Lỗi khi kiểm tra dữ liệu: {e}")
        return True
    
    for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, values_only=True):
        if all(cell is None for cell in row): #kiem tra xem value trong cell co gia tri rong khong
            return True
        else:
            return False
        
def is_recycle_bin(Ten_File): #kiem tra xem co trong thung rac khong
    # Dùng Path join để tránh lỗi escape chuỗi trên Windows
    file_path = Path("Recycle Bin") / "Ten_File"
    if file_path.exists() and file_path.is_file():
        return True

def check_file(Ten_File):
    if is_recycle_bin(Ten_File) == True: 
        return False
    elif os.path.exists(Ten_File) == False:
        return False
    elif os.path.exists(Ten_File) and is_data_none(Ten_File) == True:
        return False
    else:
        return True
    
def show_customer_information(): #show danh sach khach hang
    """Hiển thị danh sách khách hàng với error handling"""
    try:
        # Load excel file
        wb=load_workbook(filename="ThongTinKhachHang.xlsx")
        sheet=wb.active #load sheet dau tien
    except FileNotFoundError:
        print(Fore.RED + Back.BLACK + "Không tìm thấy file ThongTinKhachHang.xlsx")
        return False
    except Exception as e:
        print(Fore.RED + Back.BLACK + f"Lỗi khi mở file Excel: {e}")
        return False

    #create header from first row
    header = [cell.value for cell in sheet[1]]

    #create list of dictionaries from excel file
    data_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True): #xet tung row
        data_rows.append(dict(zip(header, row))) #ghep header + gia tri tung row
    try:
        #create dataframe from excel file
        df = pd.DataFrame(data_rows)
        print(df)
        return True
    except Exception as e:
        print(Fore.RED + Back.BLACK + f"Lỗi khi tạo DataFrame: {e}")
        return False

def create_file(): #tao file moi neu chua co 
    """Tạo file Excel mới với error handling"""
    try:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Mã KH'
        ws['B1'] = 'Họ Tên'
        ws['C1'] = 'Số ĐT'
        ws['D1'] = 'Email'
        ws['E1'] = 'Địa Chỉ'
        ws['F1'] = 'Nhóm khách hàng' #def phan loai khach hang
        ws['G1'] = 'Trạng thái' #def trang thai
        ws['H1'] = 'Lần cuối mua hàng' #dung date time
        ws['I1'] = 'Tổng tiền đã mua' #Lay data tu hoa don
        
        wb.save("ThongTinKhachHang.xlsx") #luu file
        return True
    except Exception as e:
        print(Fore.RED + Back.BLACK + f"Lỗi khi tạo file Excel: {e}")
        return False
    
   
def add_customer(): #Them khach hang
    """Thêm khách hàng mới với error handling"""
    try:
        new_customer = []
        #load excel file
        wb=load_workbook(filename="ThongTinKhachHang.xlsx")
        sheet=wb.active
    except FileNotFoundError:
        print(Fore.RED + Back.BLACK + "Không tìm thấy file ThongTinKhachHang.xlsx")
        return False
    except Exception as e:
        print(Fore.RED + Back.BLACK + f"Lỗi khi mở file Excel: {e}")
        return False
    
    #get next row
    next_row = sheet.max_row + 1
    print(Fore.CYAN + Back.BLACK + "\n ----------THÊM KHÁCH HÀNG----------")
    ma_kh = ID_kh()
    ho_ten = input("Nhập họ tên khách hàng: ")
    so_dt = str(input("Nhập số điện thoại khách hàng: "))
    while (is_dupe_data('SĐT', so_dt) == True) or (so_dt[0] != '0') or len(so_dt) != 10:
        if is_dupe_data('SĐT', so_dt) == True:
            print(Fore.RED + Back.BLACK + "Số điện thoại đã được đăng kí. Vui lòng nhập số điện thoại khác!!!")
        elif so_dt[0] != '0':
            print(Fore.RED + Back.BLACK + "Số điện thoại phải bắt đầu từ số 0")
        elif len(so_dt) != 10:
            print(Fore.RED + Back.BLACK + "Vui lòng nhập đủ 10 kí tự!!!")
        so_dt = input("Nhập số điện thoại khách hàng: ")
    dia_chi = input("Nhập địa chỉ khách hàng: ")
    email = input("Nhập email khách hàng: ")
    while is_dupe_data('Email', email) == True:
        print(Fore.RED + Back.BLACK + "Email đã được đăng kí. Vui lòng dùng Email khác!!!")
        email = input("Nhập email khách hàng: ")
    nhom_khach_hang = phan_loai()
    trang_thai = trang_thai_KH()
    lan_cuoi_mua_hang = time.time()
    # tong_tien_da_mua = data_hoa_don() 
    new_customer.append(ma_kh)
    new_customer.append(ho_ten)
    new_customer.append(so_dt)
    new_customer.append(email)
    new_customer.append(dia_chi)
    new_customer.append(nhom_khach_hang)
    new_customer.append(trang_thai)
    new_customer.append(lan_cuoi_mua_hang)
    # new_customer.append(tong_tien_da_mua)
    
    # Xử lý trường Email không bắt buộc - nếu để trống thì gán khoảng trắng
    if new_customer[3] == "" or new_customer[3].strip() == "":
        new_customer[3] = " "
    Ten_truong = ["Mã KH", "Họ Tên", "Số ĐT", "Email", "Địa Chỉ", "Nhóm khách hàng", "Trạng thái", "Lần cuối mua hàng", "Tổng tiền đã mua"]
    
    while True:
        # Chỉ kiểm tra các trường bắt buộc: Họ Tên (index 1), Số ĐT (index 2), Địa Chỉ (index 4)
        required_fields = [1, 2, 4]  # Chỉ kiểm tra Họ Tên, Số ĐT, Địa Chỉ
        validation_passed = True
        
        for i in required_fields:
            if new_customer[i] == "" or new_customer[i].strip() == "":
                print(Fore.RED + Back.BLACK + f"Trường '{Ten_truong[i]}' không được để trống. Vui lòng nhập lại.")
                validation_passed = False
                if Ten_truong[i] == "Họ Tên":
                    ho_ten = input("Nhập họ tên khách hàng: ")
                    new_customer[i] = ho_ten
                elif Ten_truong[i] == "Số ĐT":
                    so_dt = input("Nhập số điện thoại khách hàng: ")
                    new_customer[i] = so_dt
                elif Ten_truong[i] == "Địa Chỉ":
                    dia_chi = input("Nhập địa chỉ khách hàng: ")
                    new_customer[i] = dia_chi
        
        if validation_passed:
            break
    try:
        for col, value in enumerate(new_customer, start=1):
            sheet.cell(row=next_row, column=col, value=str(value)) #import value vao cell
                
        wb.save("ThongTinKhachHang.xlsx")
        print(Fore.GREEN + Back.BLACK + "Thêm khách hàng thành công.")
        return True
    except Exception as e:
        print(Fore.RED + Back.BLACK + f"Lỗi khi lưu khách hàng: {e}")
        return False
