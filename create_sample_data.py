#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script tạo dữ liệu mẫu cho hệ thống quản lý khách hàng
"""

import os
import sys
from pathlib import Path
from datetime import datetime, timedelta
import random
from defined import *

# Thêm thư mục hiện tại vào Python path
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("❌ Vui lòng cài đặt openpyxl: pip install openpyxl")
    sys.exit(1)

def create_customer_sample_data():
    """Tạo dữ liệu mẫu cho khách hàng"""
    print("📊 Tạo dữ liệu mẫu khách hàng...")
    
    # Dữ liệu mẫu khách hàng
    customers_data = [
        ["KH001", "Nguyễn Văn An", "0901234567", "an.nguyen@email.com", "123 Đường ABC, Quận 1, TP.HCM", "15,500,000đ", "15/09/2024"],
        ["KH002", "Trần Thị Bình", "0912345678", "binh.tran@email.com", "456 Đường DEF, Quận 2, TP.HCM", "8,200,000đ", "20/09/2024"],
        ["KH003", "Lê Hoàng Cường", "0923456789", "cuong.le@email.com", "789 Đường GHI, Quận 3, TP.HCM", "22,800,000đ", "10/09/2024"],
        ["KH004", "Phạm Thị Dung", "0934567890", "dung.pham@email.com", "321 Đường JKL, Quận 4, TP.HCM", "5,600,000đ", "25/09/2024"],
        ["KH005", "Hoàng Văn Em", "0945678901", "em.hoang@email.com", "654 Đường MNO, Quận 5, TP.HCM", "12,300,000đ", "18/09/2024"],
        ["KH006", "Vũ Thị Phương", "0956789012", "phuong.vu@email.com", "987 Đường PQR, Quận 6, TP.HCM", "18,900,000đ", "12/09/2024"],
        ["KH007", "Đỗ Văn Giang", "0967890123", "giang.do@email.com", "147 Đường STU, Quận 7, TP.HCM", "7,400,000đ", "22/09/2024"],
        ["KH008", "Bùi Thị Hoa", "0978901234", "hoa.bui@email.com", "258 Đường VWX, Quận 8, TP.HCM", "25,600,000đ", "08/09/2024"],
        ["KH009", "Ngô Văn Inh", "0989012345", "inh.ngo@email.com", "369 Đường YZ, Quận 9, TP.HCM", "9,800,000đ", "28/09/2024"],
        ["KH010", "Đinh Thị Kim", "0990123456", "kim.dinh@email.com", "741 Đường ABC, Quận 10, TP.HCM", "14,200,000đ", "16/09/2024"],
        ["KH011", "Lý Văn Long", "0901234568", "long.ly@email.com", "852 Đường DEF, Quận 11, TP.HCM", "6,700,000đ", "30/09/2024"],
        ["KH012", "Trương Thị Mai", "0912345679", "mai.truong@email.com", "963 Đường GHI, Quận 12, TP.HCM", "19,500,000đ", "05/09/2024"],
        ["KH013", "Phan Văn Nam", "0923456780", "nam.phan@email.com", "159 Đường JKL, Bình Thạnh, TP.HCM", "11,100,000đ", "24/09/2024"],
        ["KH014", "Cao Thị Oanh", "0934567891", "oanh.cao@email.com", "357 Đường MNO, Tân Bình, TP.HCM", "16,800,000đ", "14/09/2024"],
        ["KH015", "Võ Văn Phúc", "0945678902", "phuc.vo@email.com", "468 Đường PQR, Phú Nhuận, TP.HCM", "8,900,000đ", "26/09/2024"],
    ]
    
    try:
        # Tạo file Excel mới hoặc load file hiện có
        file_path = "Sample_Data.xlsx"
        
        if check_file(file_path) == True:
            print(f"📁 File {file_path} đã tồn tại, sẽ ghi đè...")
        elif check_file(file_path) == False:
            wb = Workbook()
            ws = wb.active
            ws.title = "KhachHang"
        
        # Tạo header
        headers = ["Mã KH", "Họ Tên", "Số ĐT", "Email", "Địa Chỉ", "Tổng tiền mua", "Ngày cuối mua"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Thêm dữ liệu
        for row_idx, customer in enumerate(customers_data, 2):
            for col_idx, value in enumerate(customer, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Lưu file
        wb.save(file_path)
        print(f"✅ Đã tạo {len(customers_data)} khách hàng mẫu trong file {file_path}")
        
        return True
        
    except Exception as e:
        print(f"❌ Lỗi khi tạo dữ liệu khách hàng: {e}")
        return False

def create_additional_users():
    """Tạo thêm một số user mẫu"""
    print("👥 Tạo thêm user mẫu...")
    
    try:
        # Import login system
        from login_system import add_new_user, load_users, save_users, users_data
        
        # Load users hiện có
        load_users()
        
        # Danh sách user mới
        new_users = [
            ("sales1", "sales123", "Nguyễn Văn Sales", 3, "READ_ONLY"),
            ("sales2", "sales456", "Trần Thị Sales", 3, "READ_ONLY"),
            ("supervisor", "super123", "Lê Văn Supervisor", 2, "READ_WRITE"),
            ("viewer", "view123", "Phạm Thị Viewer", 3, "READ_ONLY"),
        ]
        
        added_count = 0
        for username, password, full_name, level, permission in new_users:
            if username not in users_data:  # Chỉ thêm nếu chưa tồn tại
                if add_new_user(username, password, full_name, level, permission):
                    added_count += 1
                    print(f"   ✅ Đã thêm user: {username}")
                else:
                    print(f"   ❌ Không thể thêm user: {username}")
            else:
                print(f"   ⚠️  User {username} đã tồn tại, bỏ qua")
        
        print(f"✅ Đã thêm {added_count} user mới")
        return True
        
    except Exception as e:
        print(f"❌ Lỗi khi tạo user mẫu: {e}")
        return False

def main():
    """Main function"""
    print("🚀 TẠO DỮ LIỆU MẪU CHO HỆ THỐNG")
    print("=" * 50)
    
    # Tạo dữ liệu khách hàng
    if create_customer_sample_data():
        print("✅ Tạo dữ liệu khách hàng thành công!")
    else:
        print("❌ Không thể tạo dữ liệu khách hàng!")
    
    print()
    
    # Tạo thêm users
    if create_additional_users():
        print("✅ Tạo user mẫu thành công!")
    else:
        print("❌ Không thể tạo user mẫu!")
    
    print("\n🎉 HOÀN THÀNH TẠO DỮ LIỆU MẪU!")
    print("=" * 50)
    print("📋 Dữ liệu đã tạo:")
    print("   📊 15 khách hàng mẫu trong ThongTinKhachHang.xlsx")
    print("   👥 Các user mẫu trong userdatalogin.xlsx")
    print("\n🚀 Có thể chạy ứng dụng bằng: python run_app.py")

if __name__ == "__main__":
    main()
